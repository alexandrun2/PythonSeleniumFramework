import openpyxl


# pip install openpyxl //+//

n2Dictionary = {}
n2DictionaryList = []

book = openpyxl.load_workbook("C:\\Users\\alexandru.atanasoaei\\PycharmProjects1\\PythonSeleniumFramework\\TestData\\PythonDemo.xlsx")
sheet = book.active
#cell = sheet.cell(row=1, column=2) # read specific cell
#print(cell.value)

#sheet.cell(row=2, column=2).value = "AlexN2" # writing in a specific cell
#print(sheet.cell(row=2, column=2).value)

#print(sheet.max_row) ## no of rows
#print(sheet.max_column) ## no of columns

#print(sheet['A3'].value) ## value of A3 cell

for i in range(1, sheet.max_row):
        for j in range(1, sheet.max_column):
            # print(sheet.cell(row=i, column=j).value)
            n2Dictionary[sheet.cell(row=1, column=j+1).value]= sheet.cell(row=i+1, column=j+1).value

            """
            n2Dictionary.setdefault(sheet.cell(row=1, column=j + 1).value, []).append(
                sheet.cell(row=i + 1, column=j + 1).value)
            """
        n2DictionaryList.append(n2Dictionary.copy())

print(n2DictionaryList)

