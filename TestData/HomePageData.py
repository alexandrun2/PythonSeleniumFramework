import openpyxl

class HomePageData:

        test_HomePage_data = [
                            {"firstName": "Alex1", "mail": "alex1@t.com", "pass": "123", "gender": "Male"},
                            {"firstName": "Alex2", "mail": "alex2@t.com", "pass": "321", "gender": "Female"},
                            {"firstName": "Alex3", "mail": "alex3@t.com", "pass": "213", "gender": "Male"}
                            ]

        @staticmethod ## static to avoid class instantiation // to call direct the method
        def getTestDataExcel(test_case_name):
            n2Dictionary = {}

            book = openpyxl.load_workbook(
                "C:\\Users\\alexandru.atanasoaei\\PycharmProjects1\\PythonSeleniumFramework\\TestData\\PythonDemo.xlsx")
            sheet = book.active

            for i in range(1, sheet.max_row + 1):
                if sheet.cell(row=i, column=1).value == test_case_name:
                    for j in range(2, sheet.max_column + 1):
                        # print(sheet.cell(row=i, column=j).value)
                        n2Dictionary[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
                        """
                        n2Dictionary.setdefault(sheet.cell(row=1, column=j + 1).value, []).append(
                            sheet.cell(row=i + 1, column=j + 1).value)
                        """

            return[n2Dictionary] ## in order to return dictionary inside a list

        @staticmethod  ## static to avoid class instantiation // to call direct the method
        def getTestAllDataExcel():
            n2Dictionary = {}
            n2DictionaryList = []

            book = openpyxl.load_workbook(
                "C:\\Users\\alexandru.atanasoaei\\PycharmProjects1\\PythonSeleniumFramework\\TestData\\PythonDemo.xlsx")
            sheet = book.active

            for i in range(1, sheet.max_row):
                for j in range(1, sheet.max_column):
                    # print(sheet.cell(row=i, column=j).value)
                    n2Dictionary[sheet.cell(row=1, column=j + 1).value] = sheet.cell(row=i + 1, column=j + 1).value

                    """
                    n2Dictionary.setdefault(sheet.cell(row=1, column=j + 1).value, []).append(
                        sheet.cell(row=i + 1, column=j + 1).value)
                    """
                n2DictionaryList.append(n2Dictionary.copy())

            return n2DictionaryList
